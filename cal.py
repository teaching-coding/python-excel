from openpyxl import Workbook

# Step 1: Create a new Excel workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Sales Data"

# Step 2: Add sample data
data = [
    ["Product", "Units Sold", "Unit Price", "Total Sales"],
    ["Product A", 10, 5.00, None],
    ["Product B", 20, 7.50, None],
    ["Product C", 15, 6.00, None],
    ["Product D", 25, 8.00, None],
]

for row in data:
    ws.append(row)

# Step 3: Write formulas dynamically
# Calculate Total Sales for each product
for row in range(2, len(data) + 1):  # Start from the second row (excluding headers)
    ws[f"D{row}"] = f"=B{row}*C{row}"  # Units Sold * Unit Price

# Add formulas for Total Revenue and Average Revenue
ws["C7"] = "Total Revenue"
ws["D7"] = "=SUM(D2:D5)"  # Sum of Total Sales

ws["C8"] = "Average Revenue"
ws["D8"] = "=AVERAGE(D2:D5)"  # Average of Total Sales

# Step 4: Save the workbook
file_path = "sales_data_with_formulas.xlsx"
wb.save(file_path)
print(f"Workbook saved to '{file_path}'")
