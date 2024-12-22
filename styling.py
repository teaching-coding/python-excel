from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Load the existing workbook
file_name = "existing_file.xlsx"  # Replace with your file name
wb = load_workbook(file_name)

# Select the active sheet or a specific sheet by name
sheet = wb.active  # Or use wb["SheetName"]

# Add new data (if required)
new_data = [["David", 29, 58000, "Finance"],
            ["Eva", 22, 52000, "HR"]]

for row in new_data:
    sheet.append(row)

# Apply formatting to the header row
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center")

for cell in sheet[1]:  # First row is the header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment

# Apply borders to all cells
thin_border = Border(left=Side(style="thin"),
                     right=Side(style="thin"),
                     top=Side(style="thin"),
                     bottom=Side(style="thin"))

for row in sheet.iter_rows():
    for cell in row:
        cell.border = thin_border

# Adjust column widths
column_widths = [10, 5, 10, 15]  # Width for each column
for i, column_width in enumerate(column_widths, start=1):
    sheet.column_dimensions[sheet.cell(1, i).column_letter].width = column_width

# Save the changes back to the same file or a new file
wb.save("updated_file.xlsx")  # Save to a new file to preserve the original
print("Changes saved to 'updated_file.xlsx'")
